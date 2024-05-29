from openpyxl import load_workbook
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain.schema.document import Document


def handle_xlsx(file_path):
    wb = load_workbook(file_path)
    # Select the first sheet
    sheet = wb.active
    temp=""
    # Iterate over all cells in the sheet
    for row in sheet.iter_rows():
        for cell in row:
            temp += str(cell.value) + " " if cell.value is not None else ""
    text_splitter = RecursiveCharacterTextSplitter(chunk_size=2000, chunk_overlap=30)
    text = text_splitter.split_text(temp)
    docs = [Document(page_content=text[i] ,metadata={'source': file_path, 'page': i}) for i in range(len(text))]

    return docs
 

if __name__ == '__main__':
    text = handle_xlsx()
    print(text)
    